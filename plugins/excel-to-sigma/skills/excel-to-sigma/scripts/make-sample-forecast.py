#!/usr/bin/env python3
"""
Build a synthetic, self-contained Excel forecast/data-entry model that mirrors
the structure of the real Ledcor file (without using any customer data).

Shape:
  - "Forecast Entry"  : formal Table `tblForecast` at the canonical grain
                        Region x Branch x SubBranch x MonthDate x CategoryCode,
                        measure ForecastAmount. THIS is the sheet that becomes
                        a Sigma input table.
  - "Categories"      : formal Table `tblCategory` (dimension)
  - "Calendar"        : formal Table `tblCalendar` (Sept-start fiscal calendar)
  - "Forecast Summary": formula-driven rollup (SUMIFS by statement section x
                        month) -- the read/output half, rebuilt as DM metrics.

This exercises the whole excel-to-sigma surface: one formal Table at a tidy
grain (clean input-table candidate), two dimension Tables (DM relationships),
and a report sheet whose SUMIFS "evaporate" into grouped aggregates.
"""
import random
from datetime import date
from dateutil.relativedelta import relativedelta
from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter

random.seed(42)

# ---- dimensions -----------------------------------------------------------
# Region / Branch / SubBranch hierarchy
HIER = [
    ("West", "Seattle",  "SEA-North"),
    ("West", "Seattle",  "SEA-South"),
    ("West", "Portland", "PDX-Metro"),
    ("East", "Boston",   "BOS-Downtown"),
    ("East", "NewYork",  "NYC-Manhattan"),
]

# CategoryCode -> (label, statement section, sort order, typical monthly magnitude)
CATEGORIES = [
    (4000, "Product Revenue",     "Revenue",        1, ( 80000, 160000)),
    (4100, "Service Revenue",     "Revenue",        2, ( 30000,  90000)),
    (5000, "Materials",           "Contract Costs", 3, (-45000, -20000)),
    (5100, "Direct Labor",        "Contract Costs", 4, (-60000, -30000)),
    (5200, "Subcontractors",      "Contract Costs", 5, (-25000,  -8000)),
    (6000, "Admin Salaries",      "Administration", 6, (-30000, -15000)),
    (6100, "Rent & Facilities",   "Administration", 7, (-12000,  -6000)),
    (6200, "Software & IT",       "Administration", 8, ( -8000,  -3000)),
    (7000, "Corporate Allocation","Allocation",     9, (-10000,  -4000)),
]

# Sept-start fiscal year: Sep 2025 -> Aug 2026
FY_START = date(2025, 9, 1)
MONTHS = [FY_START + relativedelta(months=i) for i in range(12)]


def fiscal_bits(d):
    # Sept = fiscal month 1; fiscal year labelled by the year it ends
    fmi = (d.month - 9) % 12 + 1
    fy = d.year + 1 if d.month >= 9 else d.year
    fq = (fmi - 1) // 3 + 1
    return fmi, fy, fq


def style(ws, table_name, ref):
    t = Table(displayName=table_name, ref=ref)
    t.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2", showRowStripes=True, showColumnStripes=False)
    ws.add_table(t)


wb = Workbook()

# ---- Sheet 1: Forecast Entry (the formal Table -> input table) -------------
ws = wb.active
ws.title = "Forecast Entry"
hdr = ["Region", "Branch", "SubBranch", "MonthDate", "CategoryCode", "ForecastAmount"]
ws.append(hdr)
nrows = 0
for region, branch, sub in HIER:
    for code, label, section, order, (lo, hi) in CATEGORIES:
        for m in MONTHS:
            amt = random.randint(lo, hi)
            # light seasonality on revenue
            if section == "Revenue" and m.month in (11, 12, 1):
                amt = int(amt * 1.15)
            ws.append([region, branch, sub, m, code, amt])
            nrows += 1
for c, _ in enumerate(hdr, start=1):
    ws.column_dimensions[get_column_letter(c)].width = 16
for r in range(2, nrows + 2):
    ws.cell(r, 4).number_format = "yyyy-mm-dd"
    ws.cell(r, 6).number_format = "#,##0"
style(ws, "tblForecast", f"A1:F{nrows + 1}")

# ---- Sheet 2: Categories (dimension Table) --------------------------------
ws2 = wb.create_sheet("Categories")
ws2.append(["CategoryCode", "CategoryLabel", "StatementSection", "SortOrder"])
for code, label, section, order, _ in CATEGORIES:
    ws2.append([code, label, section, order])
for c, w in zip("ABCD", (14, 22, 18, 12)):
    ws2.column_dimensions[c].width = w
style(ws2, "tblCategory", f"A1:D{len(CATEGORIES) + 1}")

# ---- Sheet 3: Calendar (dimension Table) ----------------------------------
ws3 = wb.create_sheet("Calendar")
ws3.append(["MonthDate", "FiscalMonthIndex", "FiscalYear", "FiscalQuarter"])
for m in MONTHS:
    fmi, fy, fq = fiscal_bits(m)
    ws3.append([m, fmi, fy, fq])
for r in range(2, len(MONTHS) + 2):
    ws3.cell(r, 1).number_format = "yyyy-mm-dd"
for c, w in zip("ABCD", (14, 18, 14, 16)):
    ws3.column_dimensions[c].width = w
style(ws3, "tblCalendar", f"A1:D{len(MONTHS) + 1}")

# ---- Sheet 4: Forecast Summary (formula-driven rollup) --------------------
# Rows = statement sections; columns = the 12 fiscal months; SUMIFS rollups.
# This is the "report" sheet whose SUMIFS evaporate into grouped DM metrics.
ws4 = wb.create_sheet("Forecast Summary")
sections = ["Revenue", "Contract Costs", "Administration", "Allocation"]
# header row: blank + month labels
ws4.cell(1, 1, "Statement Section")
for j, m in enumerate(MONTHS, start=2):
    c = ws4.cell(1, j, m)
    c.number_format = "mmm-yy"
ws4.cell(1, len(MONTHS) + 2, "FY Total")

# category->section lookup lives on the Categories sheet; we SUMIFS through it
# via a helper: sum ForecastAmount where the category's section matches.
# Simplest faithful Excel idiom: SUMPRODUCT over the entry table joined to
# category sections. We approximate with SUMIFS per-category then sum by section
# in a helper block, but to keep the formula readable we use SUMPRODUCT.
# tblForecast columns: Region(A) Branch(B) SubBranch(C) MonthDate(D) CategoryCode(E) ForecastAmount(F)
fe = "'Forecast Entry'"
cat = "Categories"
ncat = len(CATEGORIES)
fr_last = nrows + 1  # last data row in Forecast Entry
for i, sec in enumerate(sections, start=2):
    ws4.cell(i, 1, sec)
    for j, m in enumerate(MONTHS, start=2):
        col = get_column_letter(j)
        # SUMPRODUCT: rows where month matches AND category's section == sec
        # category section resolved via LOOKUP of CategoryCode into Categories table
        f = (f"=SUMPRODUCT(({fe}.$D$2:$D${fr_last}={col}$1)*"
             f"(LOOKUP({fe}.$E$2:$E${fr_last},{cat}.$A$2:$A${ncat+1},{cat}.$C$2:$C${ncat+1})=$A{i})*"
             f"{fe}.$F$2:$F${fr_last})")
        cell = ws4.cell(i, j, f)
        cell.number_format = "#,##0"
    # FY total
    tcol = get_column_letter(len(MONTHS) + 2)
    ws4.cell(i, len(MONTHS) + 2, f"=SUM(B{i}:{get_column_letter(len(MONTHS)+1)}{i})").number_format = "#,##0"

# Net Contribution row
nrow = len(sections) + 2
ws4.cell(nrow, 1, "Net Contribution")
for j in range(2, len(MONTHS) + 3):
    col = get_column_letter(j)
    ws4.cell(nrow, j, f"=SUM({col}2:{col}{len(sections)+1})").number_format = "#,##0"
ws4.column_dimensions["A"].width = 20

import os
out = os.path.join(os.getcwd(), "Sample Forecast.xlsx")
wb.save(out)
print(f"wrote {out}")
print(f"tblForecast rows: {nrows}  (= {len(HIER)} subbranch x {len(CATEGORIES)} cat x {len(MONTHS)} months)")
print(f"categories: {len(CATEGORIES)}  calendar months: {len(MONTHS)}")
