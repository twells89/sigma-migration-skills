#!/usr/bin/env python3
"""
infer-canonical-formulas.py — the inference core for the EQUITY-RESEARCH archetype
(FactSet / broker house-template models: line-items down rows, YEARS across columns,
organised into statement sections — PROFIT AND LOSS, PER SHARE DATA, BALANCE SHEET, …).

It turns a "report drawn in cells" financial model into a clean two-layer plan:

  DATA LAYER   — every TYPED number in the grid (historical actual OR forecast-year
                 override — a typed cell is *always* data), long grain
                 (Period, Section, Line Item, orders, Value).
  FORMULA LAYER— for each DERIVED line item, ONE canonical formula applied to all
                 periods (the "Per Share Data done right" pattern), NOT the per-year
                 exception soup that bloats a hand build.

Why one formula, not per-year? The same row is often a hardcode in history and a
formula in the forecast (analyst overrides). We take the row's MODAL formula (after
normalising the period dimension away), let typed cells win via Coalesce in the
build step, and PARITY-GATE the result against Excel's own cached values.

Output: a research-model-plan JSON consumed by build-research-model.py. Also prints a
human report. Read-only; never writes to the source workbook.

Usage:
  infer-canonical-formulas.py <file.xlsx> [--sheet "FY results"] [--out plan.json]
                              [--first-row N] [--last-row N] [--quiet]
"""
import sys, os, json, re, math
from collections import Counter, defaultdict

try:
    import openpyxl
    from openpyxl.utils import get_column_letter, column_index_from_string
    from openpyxl.formula.tokenizer import Tokenizer, Token
except ImportError:
    sys.exit("needs openpyxl:  python3 -m venv .venv && .venv/bin/pip install openpyxl")

CELL_RE  = re.compile(r"^(\$?)([A-Z]{1,3})(\$?)(\d+)$")
RANGE_RE = re.compile(r"^(\$?[A-Z]{1,3}\$?\d+):(\$?[A-Z]{1,3}\$?\d+)$")
ERR_RE   = re.compile(r"#(REF|N/A|VALUE|DIV/0|NAME|NUM|NULL)!?", re.I)
YEAR_RE  = re.compile(r"^(19|20)\d{2}$")


# ------------------------------------------------------------------ load
def load(path):
    f = openpyxl.load_workbook(path, data_only=False)   # formulas
    v = openpyxl.load_workbook(path, data_only=True)     # cached values
    return f, v


def pick_sheet(wb, name):
    if name and name in wb.sheetnames:
        return wb[name]
    # skip FactSet cache / veryHidden; take the widest visible sheet
    cand = [ws for ws in wb.worksheets
            if ws.sheet_state == "visible" and not ws.title.startswith("__")]
    return max(cand, key=lambda w: (w.max_row or 0) * (w.max_column or 0)) if cand else wb.active


# ------------------------------------------------------------------ year axis
def detect_year_axis(ws_f, ws_v):
    """Return (header_row, [(col_idx, year_int)], first_year_col, last_year_col).
    A year column runs while the header cell is a 4-digit year (cached) or an
    increment formula (=<prev>+1). Stops at the first break (notes columns)."""
    best = None
    for r in range(1, min(ws_f.max_row, 15) + 1):
        cols = []
        for c in range(2, min(ws_f.max_column, 60) + 1):
            fv = ws_f.cell(r, c).value
            cv = ws_v.cell(r, c).value
            is_year = (isinstance(cv, (int, float)) and YEAR_RE.match(str(int(cv)))) or \
                      (isinstance(fv, str) and re.match(r"^=\s*\$?[A-Z]+\$?\d+\s*\+\s*1\s*$", fv))
            if is_year:
                cols.append(c)
            elif cols:
                break            # end of the contiguous year run
        if len(cols) > (len(best[1]) if best else 0):
            best = (r, cols)
    if not best:
        return None, [], None, None
    hr, cols = best
    axis = []
    for c in cols:
        cv = ws_v.cell(hr, c).value
        yr = int(cv) if isinstance(cv, (int, float)) and YEAR_RE.match(str(int(cv))) else None
        axis.append((c, yr))
    # fill any None years by walking the increment (=prev+1)
    for i in range(1, len(axis)):
        if axis[i][1] is None and axis[i - 1][1] is not None:
            axis[i] = (axis[i][0], axis[i - 1][1] + 1)
    return hr, axis, cols[0], cols[-1]


# ------------------------------------------------------------------ row / label map
SECTION_HINTS = ("PROFIT AND LOSS", "PER SHARE DATA", "EVALUATION", "BALANCE SHEET",
                 "CHANGE IN NET DEBT", "CASH FLOW", "COMPUTED ASSUMPTIONS", "RATIOS",
                 "INCOME STATEMENT", "VALUATION")


def is_section_header(text, has_year_data):
    if not text or has_year_data:
        return False
    t = text.strip()
    if t.upper() in SECTION_HINTS:
        return True
    # ALL-CAPS-ish, multi-word, no leading indentation, few punctuation
    letters = [ch for ch in t if ch.isalpha()]
    if len(letters) >= 3 and t == t.upper() and not t.startswith(" ") and ":" not in t:
        return True
    return False


def build_row_map(ws_f, ws_v, axis, first_row, last_row):
    """row -> {label, section, sec_order, line_order, uname, cells{col:(kind,val,formula,cached)}}
    Only rows with >=1 non-empty year cell are data rows."""
    year_cols = [c for c, _ in axis]
    rows = {}
    section, sec_order, line_order = "(none)", 0, 0
    name_seen = Counter()
    for r in range(first_row, last_row + 1):
        label = ws_f.cell(r, 1).value
        label = str(label).strip() if label is not None else ""
        # gather year cells
        cells = {}
        has_data = False
        for c in year_cols:
            fv = ws_f.cell(r, c).value
            cv = ws_v.cell(r, c).value
            if fv is None and cv is None:
                cells[c] = ("EMPTY", None, None, None); continue
            has_data = True
            if isinstance(fv, str) and fv.startswith("="):
                cells[c] = ("FORMULA", None, fv, cv)
            elif isinstance(fv, (int, float)):
                cells[c] = ("LITERAL", fv, None, cv)
            else:
                # text literal in a data row is rare; treat as label-ish, ignore for value
                cells[c] = ("TEXT", fv, None, cv)
        if is_section_header(label, has_data):
            section = label.strip(); sec_order += 1; line_order = 0
            continue
        if not has_data:
            continue    # sub-header / spacer
        line_order += 1
        # unique internal name (Sigma column). disambiguate duplicate labels.
        base = label if label else f"row{r}"
        name_seen[base] += 1
        uname = base if name_seen[base] == 1 else f"{base} ({name_seen[base]})"
        rows[r] = dict(label=label, section=section, sec_order=sec_order,
                       line_order=line_order, uname=uname, cells=cells)
    return rows


# ------------------------------------------------------------------ formula normalisation
def parse_ref(ref):
    """'B$8' -> (col_idx, row_idx, col_abs, row_abs) ; None if not a simple cell ref."""
    m = CELL_RE.match(ref)
    if not m:
        return None
    col_abs = (m.group(1) == "$")
    col_idx = column_index_from_string(m.group(2))
    row_abs = (m.group(3) == "$")
    row_idx = int(m.group(4))
    return col_idx, row_idx, col_abs, row_abs


def expand_range_rows(a, b):
    """A range like B8:B12 within one column -> list of row ints; cross-col -> None."""
    pa, pb = parse_ref(a), parse_ref(b)
    if not pa or not pb or pa[0] != pb[0]:
        return None
    return list(range(min(pa[1], pb[1]), max(pa[1], pb[1]) + 1))


def normalize_formula(formula, self_col, self_row, row_by_num):
    """Return (canonical_key, sigma_expr_template, deps:set(row), flags:set).
    canonical_key: period-relative string used for the modal vote (refs -> R<row>@<off>).
    sigma_expr_template: same but refs -> {{R<row>@<off>}} placeholders, later rendered.
    Cross-sheet refs / errors / unresolved rows set flags and make the row non-derivable
    from this cell (caller decides)."""
    flags = set()
    deps = set()
    try:
        toks = Tokenizer(formula).items
    except Exception:
        return None, None, deps, {"unparseable"}
    out = []
    for t in toks:
        val = t.value
        if t.type == Token.OPERAND and t.subtype == Token.RANGE:
            if "!" in val:                       # cross-sheet reference
                flags.add("cross_sheet"); out.append("XREF"); continue
            if ERR_RE.search(val):
                flags.add("error_ref"); out.append("ERR"); continue
            rng = RANGE_RE.match(val)
            if rng:                              # SUM(B8:B12) style
                rr = expand_range_rows(rng.group(1), rng.group(2))
                if rr is None:
                    flags.add("range_multicol"); out.append("RANGE"); continue
                parts = []
                for rw in rr:
                    if rw in row_by_num:
                        deps.add(rw); parts.append(f"R{rw}@0")
                    else:
                        flags.add("unresolved_ref"); parts.append(f"?{rw}")
                out.append("(" + "+".join(parts) + ")"); continue
            p = parse_ref(val)
            if not p:
                flags.add("weird_ref"); out.append("REF"); continue
            col, row, col_abs, row_abs = p
            if col_abs:                          # absolute column -> fixed base period
                flags.add("abs_col_ref")
            off = 0 if col_abs else (col - self_col)
            if row not in row_by_num:
                flags.add("unresolved_ref"); out.append(f"?{row}"); continue
            deps.add(row); out.append(f"R{row}@{off}")
        elif t.type == Token.OPERAND and t.subtype == Token.ERROR:
            flags.add("error_ref"); out.append("ERR")
        elif t.type == Token.OPERAND and t.subtype == Token.NUMBER:
            out.append(val)
        elif t.type == Token.OPERAND and t.subtype == Token.TEXT:
            out.append('"' + val.strip('"') + '"')
        elif t.type == Token.FUNC:
            fn = val.rstrip("(").upper()
            if fn not in ("SUM", "IF", "ROUND", "ABS", "MIN", "MAX", "INT"):
                flags.add(f"func:{fn}")
            out.append(val)
        elif t.type in (Token.PAREN, Token.SEP, Token.OP_IN, Token.OP_PRE, Token.OP_POST):
            out.append(val)
        else:
            out.append(val)
    key = "".join(out)
    return key, key, deps, flags


# ------------------------------------------------------------------ render to Sigma
UNRENDERABLE = ("XREF", "ERR", "RANGE", "REF", "?")   # tokens with no safe Sigma form


def col_id(row):
    return f"c{row}"                    # stable, ref-safe id (labels carry /,%,& — unsafe)


def _additive(key):
    """True if the canonical is a pure sum/difference of refs (Excel treats blanks as 0
    in +/-, so these can be rendered null-safe with Coalesce(ref,0) and stay LIVE)."""
    t = re.sub(r"R\d+@-?\d+", " ", key)
    t = re.sub(r"[0-9.]+", " ", t)
    t = re.sub(r"[+\-() ]", " ", t)
    return t.strip() == "" and key.strip() != ""


def render_sigma(key):
    """R<row>@<off> -> [c<row>] / Lag([c<row>],k,[Year]) ; Excel IF->If ; Sigma uses ^ for power.
    Additive formulas wrap refs in Coalesce(ref,0) to match Excel's blank-as-0 (stays live).
    Returns None if the key holds a token with no safe Sigma rendering (cross-sheet, error, …)."""
    if any(tok in key for tok in UNRENDERABLE):
        return None
    add = _additive(key)

    def repl(m):
        row = int(m.group(1)); off = int(m.group(2))
        nm = col_id(row)
        base = f"[{nm}]" if off == 0 else (
            f"Lag([{nm}], {-off}, [Year])" if off < 0 else f"Lead([{nm}], {off}, [Year])")
        return f"Coalesce({base}, 0)" if add else base
    s = re.sub(r"R(\d+)@(-?\d+)", repl, key)
    s = re.sub(r"\bIF\(", "If(", s)
    s = re.sub(r"\bSUM\(", "Sum(", s)
    s = re.sub(r"([0-9.]+)\s*%", r"(\1 * 0.01)", s)         # Excel 101% -> (101*0.01)
    # Sigma rejects a unary '+' prefix (Excel '=+B8+B11'): drop '+' at start / after '(' or ','
    for _ in range(3):
        s = re.sub(r"(^|[(,])\s*\+", r"\1", s.strip())
    return s if _safe_sigma(s) else None                     # unsupported -> carry as data


_ALLOWED_FUNCS = ("If", "Coalesce", "Lag", "Lead", "Sum", "Min", "Max", "Round", "Abs", "Power")


def _safe_sigma(s):
    """True iff the formula uses only constructs we can build (refs, arithmetic, allowed
    funcs). Anything else (& concat, text, unknown functions) -> carry as cached data."""
    t = re.sub(r"\[c\d+\]", " ", s)                          # column refs
    for fn in _ALLOWED_FUNCS:
        t = re.sub(rf"\b{fn}\(", " ", t)
    t = re.sub(r"[0-9]+(\.[0-9]+)?", " ", t)                 # numbers
    t = re.sub(r"[\s+\-*/^(),.<>=]", " ", t)                 # operators / punctuation
    return t.strip() == ""


# ------------------------------------------------------------------ parity eval
def eval_key(key, self_col, cached):
    """Evaluate a canonical key against the cached grid for one period column.
    cached[(row,col)] -> float. Returns float or None (unevaluatable / missing)."""
    if any(tok in key for tok in ("XREF", "ERR", "REF", "RANGE", "?")):
        return None
    key = re.sub(r"([0-9.]+)\s*%", r"(\1*0.01)", key)        # Excel percent literal
    if re.search(r"\b(IF|MIN|MAX|ROUND|ABS|INT)\(", key, re.I):
        return None            # skip strict eval on branching/functions (v1)

    def repl(m):
        row = int(m.group(1)); off = int(m.group(2))
        v = cached.get((row, self_col + off))
        if not isinstance(v, (int, float)):
            raise ValueError("missing")
        return repr(float(v))
    try:
        expr = re.sub(r"R(\d+)@(-?\d+)", repl, key)
        expr = expr.replace("^", "**")
        if not re.fullmatch(r"[0-9eE.+\-*/() ]*", expr):
            return None
        return eval(expr, {"__builtins__": {}}, {})   # sanitised: digits/operators only
    except Exception:
        return None


def approx(a, b, abs_tol=0.01, rel_tol=0.001):
    if a is None or b is None:
        return False
    if abs(a - b) <= abs_tol:
        return True
    denom = max(abs(a), abs(b), 1e-9)
    return abs(a - b) / denom <= rel_tol


def _eval_rform(key, pos, computed):
    """Evaluate a canonical KEY (R<row>@<off> form) at year-position `pos`, using the
    already-computed grid, with Sigma's null-propagation (any None operand -> None)."""
    if not key or any(t in key for t in ("XREF", "ERR", "REF", "RANGE", "?")):
        return None
    if re.search(r"\b(IF|MIN|MAX|ROUND|ABS|INT)\(", key, re.I):
        return None
    key = re.sub(r"([0-9.]+)\s*%", r"(\1*0.01)", key)
    add = _additive(key)                                   # additive -> blanks are 0 (like Excel)
    for r, o in re.findall(r"R(\d+)@(-?\d+)", key):
        v = computed.get(int(r), {}).get(pos + int(o))
        if not isinstance(v, (int, float)) and not add:
            return None                                    # null propagates (non-additive)

    def sub(m):
        v = computed.get(int(m.group(1)), {}).get(pos + int(m.group(2)))
        return repr(float(v)) if isinstance(v, (int, float)) else ("0.0" if add else "None")
    expr = re.sub(r"R(\d+)@(-?\d+)", sub, key)
    expr = expr.replace("^", "**")
    if not re.fullmatch(r"[0-9eE.+\-*/() ]*", expr):
        return None
    try:
        return eval(expr, {"__builtins__": {}}, {})       # sanitised: digits/operators only
    except Exception:
        return None


def simulate_and_freeze(plan_lines, rows, axis, cached):
    """Compute the model the way Sigma will (Coalesce(typed, canonical), topological,
    null-propagating), compare to Excel cached, and freeze cached values where the live
    computation doesn't reproduce them. Iterates because a freeze fixes downstream cells.
    Returns (live_cells, frozen_cells, [rows with genuine numeric disagreement])."""
    year_cols = [c for c, _ in axis]
    col_year = dict(axis)
    pos_of = {c: i for i, c in enumerate(year_cols)}
    by_row = {p["row"]: p for p in plan_lines}
    derived = [p for p in plan_lines if p["kind"] in ("derived", "ratio") and p.get("canonical_key")]

    # topological order by SAME-PERIOD deps (offset 0); prior/future handled by year order
    graph = {p["row"]: {int(m.group(1)) for m in re.finditer(r"R(\d+)@0", p["canonical_key"])}
                        & set(by_row) for p in derived}
    order, seen = [], set()

    def visit(r, stack):
        if r in seen or r not in graph:
            return
        for d in graph.get(r, ()):
            if d not in stack:
                visit(d, stack | {r})
        seen.add(r); order.append(r)
    for p in derived:
        visit(p["row"], set())

    mism_rows = set()
    disagree = defaultdict(int)
    for _ in range(6):                                     # converge (freezes are monotonic)
        disagree = defaultdict(int)                        # recount each pass; keep the last
        computed = defaultdict(dict)
        # seed inputs + already-frozen/override typed values
        for p in plan_lines:
            for c in year_cols:
                yr = col_year[c]
                if yr in p["typed"]:
                    computed[p["row"]][pos_of[c]] = float(p["typed"][yr])
        # compute derived in (year, topo) order
        new_freeze = 0
        for pos in range(len(year_cols)):
            for r in order:
                p = by_row[r]
                yr = col_year[year_cols[pos]]
                if yr in p["typed"]:                       # Coalesce: typed/override/frozen wins
                    continue
                computed[r][pos] = _eval_rform(p["canonical_key"], pos, computed)
        # compare to cached; freeze where live result != Excel value
        for p in derived:
            for c in year_cols:
                cv = cached.get((p["row"], c))
                if not isinstance(cv, (int, float)):
                    continue
                if col_year[c] in p["typed"]:
                    continue                               # already frozen/typed
                got = computed[p["row"]].get(pos_of[c])
                if got is None or not approx(got, cv):
                    p["typed"][col_year[c]] = round(cv, 6)  # freeze
                    new_freeze += 1
                    if got is not None:                    # a genuine numeric disagreement
                        disagree[p["row"]] += 1
        if new_freeze == 0:
            break

    # flag NEEDS_REVIEW only where the one canonical is WIDELY wrong (>25% of formula years),
    # not for the odd structural-override year (those are just frozen, and correct).
    for p in derived:
        fcells = sum(1 for c in year_cols
                     if rows[p["row"]]["cells"][c][0] == "FORMULA"
                     and isinstance(rows[p["row"]]["cells"][c][3], (int, float)))
        if fcells and disagree[p["row"]] / fcells > 0.25:
            mism_rows.add(p["row"])

    # final tally: for every formula cell with a cached value, live vs frozen
    live = frozen = 0
    for p in derived:
        for c in year_cols:
            k, lit, fml, cv = rows[p["row"]]["cells"][c]
            if k != "FORMULA" or not isinstance(cv, (int, float)):
                continue
            if col_year[c] in p["typed"]:
                frozen += 1
            else:
                live += 1
    for r in mism_rows:
        by_row[r]["reasons"].append("parity: canonical disagreed with Excel in >=1 year "
                                    "-> Excel value frozen (carried)")
    return live, frozen, [by_row[r] for r in mism_rows]


# ------------------------------------------------------------------ main inference
def infer(path, sheet=None, first_row=None, last_row=None):
    wb_f, wb_v = load(path)
    ws_f = pick_sheet(wb_f, sheet)
    ws_v = wb_v[ws_f.title]
    hr, axis, fc, lc = detect_year_axis(ws_f, ws_v)
    if not axis:
        raise ValueError(f"could not detect a year axis on sheet {ws_f.title!r}")
    first_row = first_row or (hr + 1)
    last_row = last_row or ws_f.max_row
    rows = build_row_map(ws_f, ws_v, axis, first_row, last_row)
    row_by_num = rows
    year_cols = [c for c, _ in axis]
    col_year = dict(axis)

    # cached grid for parity
    cached = {}
    for r, rd in rows.items():
        for c, (kind, lit, fml, cv) in rd["cells"].items():
            if isinstance(cv, (int, float)):
                cached[(r, c)] = float(cv)

    # per-row inference
    plan_lines = []
    stats = Counter()
    for r in sorted(rows):
        rd = rows[r]
        keys = Counter()             # canonical key -> count (formula cells only)
        key_years = defaultdict(list)
        all_deps = set()
        row_flags = set()
        typed = {}                   # year -> value (literals + cross-sheet cached = data)
        n_formula = n_literal = n_xsheet = n_empty = 0
        for c in year_cols:
            kind, lit, fml, cv = rd["cells"][c]
            yr = col_year[c]
            if kind == "EMPTY":
                n_empty += 1; continue
            if kind in ("LITERAL", "TEXT"):
                n_literal += 1
                if isinstance(lit, (int, float)):
                    typed[yr] = float(lit)
                continue
            # FORMULA
            key, _, deps, flags = normalize_formula(fml, c, r, row_by_num)
            if "cross_sheet" in flags:
                n_xsheet += 1
                if isinstance(cv, (int, float)):
                    typed[yr] = float(cv)        # linked value -> treat as data input
                row_flags.add("has_cross_sheet"); continue
            if key is None or {"error_ref", "unparseable"} & flags:
                row_flags |= flags; continue     # broken cell: ignore for voting
            n_formula += 1
            keys[key] += 1; key_years[key].append(yr)
            all_deps |= deps
            row_flags |= {f for f in flags if f.startswith("func:") or f in
                          ("abs_col_ref", "unresolved_ref", "range_multicol")}

        # decide kind
        if n_formula == 0:
            plan_lines.append(dict(row=r, col_id=col_id(r), uname=rd["uname"], label=rd["label"],
                section=rd["section"], sec_order=rd["sec_order"], line_order=rd["line_order"],
                kind="input", typed=typed, deps=[], canonical=None, status="OK", reasons=[]))
            continue

        modal, modal_n = keys.most_common(1)[0]
        reasons = []
        if len(keys) > 1:
            reasons.append(f"disputed:{len(keys)} formula variants "
                           f"(modal used, minority years {sorted(sum([key_years[k] for k in keys if k!=modal],[]))})")
        for f in row_flags:
            if f.startswith("func:"):
                reasons.append(f"contains {f.split(':')[1]}()")
            elif f in ("unresolved_ref", "abs_col_ref", "range_multicol"):
                reasons.append(f)
        canonical_sigma = render_sigma(modal)
        if canonical_sigma is None:
            # no safe Sigma form (cross-sheet/error/multicol range) -> carry cached values as data
            plan_lines.append(dict(row=r, col_id=col_id(r), uname=rd["uname"], label=rd["label"],
                section=rd["section"], sec_order=rd["sec_order"], line_order=rd["line_order"],
                kind="input", typed=typed, deps=[], canonical=None, status="OK",
                reasons=reasons + ["unrenderable formula -> carried as typed data"]))
            continue
        kind = "ratio" if _looks_ratio(modal, rd["label"]) else "derived"
        plan_lines.append(dict(row=r, col_id=col_id(r), uname=rd["uname"], label=rd["label"],
            section=rd["section"], sec_order=rd["sec_order"], line_order=rd["line_order"],
            kind=kind, typed=typed, deps=sorted(all_deps), canonical=canonical_sigma,
            canonical_key=modal, status="OK", reasons=reasons))

    # DAG: break analyst "plug cycles" by carrying the most-typed member as data
    by_row = {p["row"]: p for p in plan_lines}
    carried = break_cycles(plan_lines, by_row)
    # freeze EVERY cached value on any input/carried line (formula cells included) so the data
    # page shows all Excel values — and do it BEFORE the simulation so downstream derived lines
    # can read these carried values and stay live.
    for p in plan_lines:
        if p["kind"] == "input":
            for c in year_cols:
                cv = cached.get((p["row"], c))
                if isinstance(cv, (int, float)):
                    p["typed"].setdefault(col_year[c], round(float(cv), 6))

    # parity gate: simulate the WHOLE model exactly as Sigma will evaluate it (topological,
    # null-propagating), then freeze the Excel cached value into any cell the live formulas
    # don't reproduce. Guarantees displayed parity = 100% (live where verified, cached else).
    parity_live, parity_frozen, mism_rows = simulate_and_freeze(plan_lines, rows, axis, cached)
    for p in mism_rows:
        p["status"] = "NEEDS_REVIEW"

    # final accounting (after demotion + parity)
    stats = Counter()
    for p in plan_lines:
        stats[p["kind"]] += 1
        if "carried as typed data" in " ".join(p.get("reasons", [])):
            stats["carried_plug"] += 1
        if p["status"] == "NEEDS_REVIEW":
            stats["needs_review"] += 1

    anchors = _anchors(plan_lines)
    # ground-truth anchor values per year (from Excel cached) so the builder can assert live parity
    expected = []
    for nm, a in anchors.items():
        for c in year_cols:
            v = cached.get((a["row"], c))
            if isinstance(v, (int, float)):
                expected.append({"anchor": nm, "col_id": a["col_id"], "year": col_year[c],
                                 "value": round(v, 4)})
    plan = dict(
        source=os.path.basename(path), sheet=ws_f.title,
        periods=[{"year": y, "col": get_column_letter(c)} for c, y in axis],
        lines=plan_lines,
        anchors=anchors,
        expected_anchors=expected,
        summary=dict(stats),
        parity=dict(live=parity_live, frozen=parity_frozen),
    )
    return plan


def _looks_ratio(key, label):
    lab = (label or "").lower().strip()
    if lab.startswith("%") or lab.startswith(". %") or "% " in lab[:6] or "rate" in lab:
        return True
    # a single division of two refs, no + or -
    body = re.sub(r"R\d+@-?\d+", "R", key)
    return bool(re.fullmatch(r"R/R", body.replace(" ", ""))) or body.strip().endswith("-1")


def _sccs(graph):
    """Tarjan strongly-connected components over {node: [deps]}."""
    index = {}; low = {}; onstack = {}; stack = []; out = []; counter = [0]

    def strong(v):
        index[v] = low[v] = counter[0]; counter[0] += 1
        stack.append(v); onstack[v] = True
        for w in graph.get(v, []):
            if w not in index:
                strong(w); low[v] = min(low[v], low[w])
            elif onstack.get(w):
                low[v] = min(low[v], index[w])
        if low[v] == index[v]:
            comp = []
            while True:
                w = stack.pop(); onstack[w] = False; comp.append(w)
                if w == v:
                    break
            out.append(comp)

    for v in list(graph):
        if v not in index:
            strong(v)
    return out


def _same_period_refs(canonical):
    """Rows referenced at the SAME period by a rendered canonical (the edges Sigma sees).
    Refs inside Lag(...)/Lead(...) are prior/future period -> NOT static-cycle edges."""
    if not canonical:
        return set()
    stripped = re.sub(r"(Lag|Lead)\([^)]*\)", " ", canonical)   # drop prior/future-period refs
    return {int(m.group(1)) for m in re.finditer(r"\[c(\d+)\]", stripped)}


def break_cycles(plan_lines, by_row):
    """Break analyst plug-cycles (row A computes from B in some years, B from A in
    others — e.g. history types COGS→Gross profit, forecast types Gross profit→COGS) AND
    same-period self-references. A static Sigma column DAG can't hold either, so demote the
    MOST-TYPED cycle member to carried data ('start from hard-coded values'). The dep graph
    is read from the RENDERED canonical so it matches exactly what Sigma will evaluate.
    Returns the set of demoted (carried) rows."""
    carried = set()
    for _ in range(500):                                        # iterate until acyclic
        graph = {p["row"]: [d for d in _same_period_refs(p.get("canonical"))
                            if d in by_row and by_row[d]["kind"] in ("derived", "ratio")
                            and by_row[d]["row"] not in carried]
                 for p in plan_lines
                 if p["kind"] in ("derived", "ratio") and p["row"] not in carried}
        cyclic = [comp for comp in _sccs(graph)
                  if len(comp) > 1 or (comp[0] in graph.get(comp[0], []))]   # incl self-loop
        if not cyclic:
            break
        for comp in cyclic:
            indeg = Counter(d for r in comp for d in graph.get(r, []) if d in comp)
            victim = max(comp, key=lambda r: (len(by_row[r]["typed"]), indeg[r], -r))
            p = by_row[victim]
            p["kind"] = "input"; p["canonical"] = None; p.pop("canonical_key", None)
            p["reasons"].append("plug-cycle: carried as typed data (not recomputed live)")
            carried.add(victim)
    for p in plan_lines:
        if any("unresolved_ref" in x for x in p.get("reasons", [])) and p["kind"] != "input":
            p["status"] = "NEEDS_REVIEW"
    return carried


def _anchors(plan_lines):
    want = {"turnover": "Turnover", "ebit ": "EBIT", "= ebit": "EBIT",
            "net income": "Net income", "net attributable": "Net attributable",
            "eps reported": "EPS reported", "eps adjusted": "EPS adjusted"}
    found = {}
    for p in plan_lines:
        low = (p["label"] or "").lower()
        for k, nm in want.items():
            if k in low and nm not in found:
                found[nm] = {"col_id": p["col_id"], "label": p["label"], "row": p["row"]}
    return found


# ------------------------------------------------------------------ report
def report(plan):
    print(f"# Canonical-formula inference — {plan['source']}  (sheet {plan['sheet']!r})")
    yrs = [str(p["year"]) for p in plan["periods"]]
    print(f"\nPeriods ({len(yrs)}): {yrs[0]}…{yrs[-1]}")
    s = plan["summary"]
    print(f"\nLine items: input={s.get('input',0)}  derived={s.get('derived',0)}  "
          f"ratio={s.get('ratio',0)}  (of which carried plug-cycles={s.get('carried_plug',0)})  "
          f"NEEDS_REVIEW={s.get('needs_review',0)}")
    par = plan["parity"]
    tot = par["live"] + par["frozen"]
    rate = (100 * par["live"] / tot) if tot else 0
    print(f"Formula cells: {par['live']}/{tot} live-verified ({rate:.1f}%); "
          f"{par['frozen']} frozen to Excel cached value (carried). Displayed parity = 100%.")
    print(f"Anchors resolved: {plan['anchors']}")
    rev = [p for p in plan["lines"] if p["status"] == "NEEDS_REVIEW"]
    if rev:
        print(f"\nNEEDS_REVIEW ({len(rev)}):")
        for p in rev[:25]:
            print(f"  [{p['section']}] {p['label']!r}: {'; '.join(p['reasons'])}")
    print("\nSample DERIVED canonicals:")
    shown = 0
    for p in plan["lines"]:
        if p["kind"] == "derived" and p["status"] == "OK" and p.get("canonical") and shown < 14:
            print(f"  {p['label'][:40]:40s} = {p['canonical']}")
            shown += 1


def main():
    a = [x for x in sys.argv[1:] if not x.startswith("--")]
    opt = {x.split("=")[0]: (x.split("=", 1)[1] if "=" in x else True) for x in sys.argv[1:] if x.startswith("--")}
    if not a:
        sys.exit(__doc__)
    plan = infer(a[0], sheet=opt.get("--sheet"),
                 first_row=int(opt["--first-row"]) if "--first-row" in opt else None,
                 last_row=int(opt["--last-row"]) if "--last-row" in opt else None)
    if "--quiet" not in opt:
        report(plan)
    out = opt.get("--out")
    if out and out is not True:
        json.dump(plan, open(out, "w"), indent=1)
        print(f"\nplan written: {out}")


if __name__ == "__main__":
    main()
