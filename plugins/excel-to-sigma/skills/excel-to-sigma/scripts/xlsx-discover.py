#!/usr/bin/env python3
"""Phase 0 — discover an .xlsx workbook.

Inventories formal Tables (the migration's Rosetta Stone), pivots/charts, and a
formula census (flagging untranslatable / report-style sheets). Prints a summary
and, for each formal Table, its columns + inferred grain.

Usage:
    python xlsx-discover.py "Sample Forecast.xlsx"
"""
import sys
import re
from collections import Counter
from openpyxl import load_workbook

# functions we translate today (see refs/excel-translation.md)
TRANSLATABLE = {
    "IF", "IFS", "SWITCH", "IFERROR", "AND", "OR", "NOT",
    "SUM", "SUMIF", "SUMIFS", "COUNT", "COUNTA", "COUNTIF", "COUNTIFS",
    "AVERAGE", "AVERAGEIF", "AVERAGEIFS", "MIN", "MAX", "MEDIAN",
    "VLOOKUP", "XLOOKUP", "HLOOKUP", "INDEX", "MATCH", "LOOKUP",
    "LEFT", "RIGHT", "MID", "LEN", "CONCAT", "CONCATENATE", "TRIM",
    "UPPER", "LOWER", "SUBSTITUTE", "FIND", "TEXT",
    "TODAY", "NOW", "YEAR", "MONTH", "DAY", "DATEDIF", "EDATE",
    "EOMONTH", "WEEKDAY", "DATE",
    "ROUND", "ROUNDUP", "ROUNDDOWN", "CEILING", "FLOOR", "ABS", "POWER", "MOD",
    "SUMPRODUCT",
}
# fail-loud: these usually mark a "report drawn in cells" needing manual rebuild
VOLATILE = {"OFFSET", "INDIRECT", "INDEX_VOLATILE"}

FUNC_RE = re.compile(r"\b([A-Z][A-Z0-9_.]+)\s*\(")
PK_RE = re.compile(r"(^id|_id|id|code|key|num)$", re.I)
# headers that look like a spread dimension (a wide cross-tab axis)
WIDE_HDR_RE = re.compile(
    r"^(\d{4}([-_/ ]?\d{1,2})?|\d{1,2}[-_/ ]\d{2,4}|"
    r"jan|feb|mar|apr|may|jun|jul|aug|sep|oct|nov|dec|q[1-4]|fy\d+|\d+)$", re.I)


def classify(headers, body_rows, nrows, referenced_by_others):
    """Return (route, why) — entered → input table; derived → read-only DM;
    reference → maintained-list (input table if users maintain it)."""
    # formula ratio over non-empty body cells
    filled = formulas = 0
    for r in body_rows:
        for c in r:
            v = c.value
            if v is None or v == "":
                continue
            filled += 1
            if isinstance(v, str) and v.startswith("="):
                formulas += 1
    fr = (formulas / filled) if filled else 0.0
    pk = any(PK_RE.search(str(h).strip()) for h in headers if h)
    # wide cross-tab: many headers that look like a spread axis
    wide = sum(1 for h in headers if h and WIDE_HDR_RE.match(str(h).strip()))
    wide_flag = wide >= 4

    if fr >= 0.3:
        return "DERIVED", f"{fr:.0%} of cells are formulas → read-only DM element + metrics", wide_flag
    if nrows <= 30 and pk and referenced_by_others:
        return "REFERENCE", "small + PK + referenced by other sheets → dim (input table only if users maintain it; else read-only)", wide_flag
    return "ENTERED", "typed values, no inbound formulas → input table (editable)", wide_flag


SHEET_REF_RE = re.compile(r"(?:'([^']+)'|([A-Za-z0-9_]+))!\$?[A-Z]+\$?\d+")
GROWTH_RE = re.compile(r"^=\s*\$?[A-Z]+\$?\d+\s*\*\s*\(\s*1\s*[+\-]", re.I)   # prior*(1±rate)
FLAT_RE   = re.compile(r"^=\s*\$?[A-Z]+\$?\d+\s*$")                          # =single cell
SUMIFS_RE = re.compile(r"\bSUMIFS?\b|\bSUMPRODUCT\b", re.I)
ARITH_RE  = re.compile(r"[A-Z]+\$?\d+\s*[-+*/]\s*\$?[A-Z]+\$?\d+|^=IF\(", re.I)

def classify_report_cells(wb):
    """Cell-level model map for 'report drawn in cells' sheets (no formal Table).
    Returns per-sheet counts by class + the detected source/driver sheets."""
    table_sheets = {ws.title for ws in wb.worksheets if len(ws.tables)}
    sheetmap = {}          # sheet -> Counter of classes
    source_sheets = Counter()   # tabs feeding SUMIFS (actuals source)
    driver_sheets = Counter()   # tabs referenced by growth (1+rate) formulas
    for ws in wb.worksheets:
        if ws.title in table_sheets:
            continue
        counts = Counter()
        # which rows contain a formula (to spot manual literals sitting among formulas)
        rows_with_formula = set()
        numeric_cells = []
        for row in ws.iter_rows():
            for c in row:
                v = c.value
                if isinstance(v, str) and v.startswith("="):
                    rows_with_formula.add(c.row)
                elif isinstance(v, (int, float)) and c.column > 1:
                    numeric_cells.append(c)
        for row in ws.iter_rows():
            for c in row:
                v = c.value
                if not (isinstance(v, str) and v.startswith("=")):
                    continue
                if SUMIFS_RE.search(v):
                    counts["ACTUAL"] += 1
                    for m in SHEET_REF_RE.finditer(v):
                        source_sheets[m.group(1) or m.group(2)] += 1
                elif GROWTH_RE.match(v):
                    counts["DRIVER_GROWTH"] += 1
                    for m in SHEET_REF_RE.finditer(v):
                        driver_sheets[m.group(1) or m.group(2)] += 1
                elif FLAT_RE.match(v):
                    counts["FLAT"] += 1
                elif ARITH_RE.search(v):
                    counts["DERIVED"] += 1
                else:
                    counts["OTHER_FORMULA"] += 1
        # manual inputs = numeric literals sitting in rows that are otherwise formula-driven
        manual = sum(1 for c in numeric_cells if c.row in rows_with_formula)
        if manual:
            counts["MANUAL"] = manual
        if counts:
            sheetmap[ws.title] = counts
    return sheetmap, source_sheets, driver_sheets


def main(path):
    wb = load_workbook(path, data_only=False)
    print(f"# Discovery: {path}\n")
    print(f"Sheets ({len(wb.sheetnames)}): {', '.join(wb.sheetnames)}\n")

    # all formula strings (with their sheet) — for cross-sheet reference detection
    all_formulas = []
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for c in row:
                if isinstance(c.value, str) and c.value.startswith("="):
                    all_formulas.append((ws.title, c.value))

    # ---- formal Tables + routing ----
    tables = []
    for ws in wb.worksheets:
        for name in ws.tables:
            ref = ws.tables[name].ref
            cells = ws[ref]
            headers = [c.value for c in cells[0]] if cells else []
            body = cells[1:] if cells else []
            nrows = len(body)
            referenced = any(s != ws.title and (ws.title in f or name in f)
                             for s, f in all_formulas)
            route, why, wide = classify(headers, body, nrows, referenced)
            tables.append((name, ws.title, ref, headers, nrows, route, why, wide))

    print(f"## Formal Tables ({len(tables)}) — with routing")
    routes = Counter()
    for name, sheet, ref, headers, nrows, route, why, wide in tables:
        routes[route] += 1
        tag = {"ENTERED": "✏️  ENTERED", "DERIVED": "🔒 DERIVED", "REFERENCE": "📒 REFERENCE"}[route]
        print(f"\n  • {name}  (sheet '{sheet}', {ref}, ~{nrows} rows)  →  {tag}")
        print(f"    columns: {', '.join(str(h) for h in headers)}")
        print(f"    routing: {why}")
        if wide:
            print(f"    ⚠️ wide cross-tab detected → UNPIVOT (Sigma transpose / SQL unpivot) to a tidy grain first")
    if not tables:
        print("  (none — workbook is likely 'report drawn in cells'; expect manual rebuild)")
    else:
        print(f"\n  routing summary: " + ", ".join(f"{n} {r.lower()}" for r, n in routes.items()))
        print("  → ENTERED/REFERENCE-maintained tables become input tables (preserve data entry);")
        print("    DERIVED tables become read-only DM elements + metrics. See SKILL.md 'Preserve the inputs'.")

    # ---- pivots / charts ----
    n_charts = sum(len(getattr(ws, "_charts", [])) for ws in wb.worksheets)
    n_pivots = sum(len(getattr(ws, "_pivots", [])) for ws in wb.worksheets)
    print(f"\n## Visuals: {n_charts} chart(s), {n_pivots} pivot table(s)")

    # ---- formula census ----
    func_counts = Counter()
    formula_cells = 0
    report_sheets = Counter()
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for c in row:
                if isinstance(c.value, str) and c.value.startswith("="):
                    formula_cells += 1
                    for m in FUNC_RE.findall(c.value):
                        func_counts[m] += 1
                        if m in VOLATILE:
                            report_sheets[ws.title] += 1
    print(f"\n## Formula census: {formula_cells} formula cells, "
          f"{len(func_counts)} distinct functions")
    untranslatable = {f: n for f, n in func_counts.items()
                      if f not in TRANSLATABLE and f in VOLATILE}
    unknown = {f: n for f, n in func_counts.items()
               if f not in TRANSLATABLE and f not in VOLATILE}
    top = func_counts.most_common(15)
    print("  top functions:", ", ".join(f"{f}×{n}" for f, n in top))
    if untranslatable:
        print(f"  ⚠️ volatile/untranslatable (manual rebuild): "
              + ", ".join(f"{f}×{n}" for f, n in untranslatable.items()))
        print(f"     report-style sheets: {', '.join(report_sheets)}")
    if unknown:
        print(f"  ❓ not yet in the rules table (verify): "
              + ", ".join(f"{f}×{n}" for f, n in sorted(unknown.items())))

    # ---- derived report sheets (formula-driven, no formal table) ----
    table_sheets = {ws.title for ws in wb.worksheets if len(ws.tables)}
    per_sheet = Counter(s for s, _ in all_formulas)
    derived = [s for s, n in per_sheet.items() if n and s not in table_sheets]
    if derived:
        print(f"\n## Derived report sheets (formula-driven, no formal table): "
              f"{', '.join(derived)}")
        print("  🔒 read-only — rebuild as DM metrics + workbook elements (the SUMIFS/"
              "rollups become grouped aggregates), not input tables.")

    # ---- cell-level MODEL MAP (for report-drawn-in-cells models; see refs/model-taxonomy.md) ----
    sheetmap, source_sheets, driver_sheets = classify_report_cells(wb)
    if sheetmap:
        print("\n## Model map (cell-level) — for report sheets drawn in cells")
        ICON = {"ACTUAL": "📥 ACTUAL (live source)", "DRIVER_GROWTH": "📈 DRIVER-grown",
                "MANUAL": "✏️  MANUAL input", "DERIVED": "🧮 DERIVED", "FLAT": "➡️  FLAT (carry)",
                "OTHER_FORMULA": "· other"}
        for sheet, counts in sheetmap.items():
            parts = [f"{ICON.get(k,k)}={v}" for k, v in counts.most_common()]
            print(f"  • {sheet}: " + ", ".join(parts))
        if source_sheets:
            print(f"  → actuals source (SUMIFS targets): {', '.join(source_sheets)}  "
                  f"= live read-only DM element")
        if driver_sheets:
            print(f"  → driver/assumptions sheet (1±rate refs): {', '.join(driver_sheets)}  "
                  f"= editable rates table / controls")
        print("  Architecture: union [live ACTUAL] + [forecast]; forecast = DRIVER-grown "
              "(base×(1+rate)^n, rate from drivers) + MANUAL (editable input table) + FLAT. "
              "Subtotals/margins = DERIVED workbook calcs. See refs/forecast-recipes.md.")
        print("  ASK INTENT before building (read-only report / editable plan / live what-if) "
              "→ determines static vs input-table vs driver-override. See refs/model-taxonomy.md.")

    print("\nNext: classify intent → build per refs/forecast-recipes.md; ENTERED → input-table "
          "builder; DERIVED → read-only DM; wide → unpivot first; macros → macro-classify.py.")


if __name__ == "__main__":
    if len(sys.argv) < 2:
        sys.exit("usage: xlsx-discover.py <file.xlsx>")
    main(sys.argv[1])
