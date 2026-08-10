#!/usr/bin/env python3
"""
make-sample-research-model.py — generate a SYNTHETIC broker equity-research model
(no customer data) shaped like a FactSet-style broker house-template file, for use as the
excel-to-sigma golden fixture + QUICKSTART example.

Fake issuer "Acme Widgets NV", fake broker "Zenith Securities". Sheet "FY results":
line-items down rows, YEARS across columns, sections PROFIT AND LOSS / PER SHARE DATA /
BALANCE SHEET. Deliberately reproduces the hard cases so the converter is exercised:
  - MIXED rows (hardcoded history, formula forecast),
  - a plug-cycle (COGS types in history, Gross profit types in forecast),
  - a `%` literal (Excel 105%),
  - a same-period self-reference and a `#REF!` broken cell (escape-hatch test),
  - ratio rows (% margin, % growth) and a prior-year %change via a cross-year ref.

Writes scripts/Sample Research Model.xlsx.
"""
import os, re, zipfile, shutil
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

YEARS = list(range(2018, 2028))          # 10 years; 2018-2023 actual, 2024-2027 estimate
FIRST_COL = 2                            # B


def col(i):
    return get_column_letter(FIRST_COL + i)


def build():
    wb = Workbook()
    ws = wb.active; ws.title = "FY results"
    ws["A3"] = "Acme Widgets NV"
    ws["A4"] = "12M to Dec"
    # year header row 5: first two hardcoded, rest = prev+1 (as the real files do)
    ws[f"{col(0)}5"] = YEARS[0]; ws[f"{col(1)}5"] = YEARS[1]
    for i in range(2, len(YEARS)):
        ws[f"{col(i)}5"] = f"={col(i-1)}5+1"

    def put(r, label, values):
        ws[f"A{r}"] = label
        for i, v in enumerate(values):
            if v is None:
                continue
            ws[f"{col(i)}{r}"] = v

    # ---- PROFIT AND LOSS ----
    ws["A7"] = "PROFIT AND LOSS"
    # Revenue: actuals hardcoded, forecast grows 5% (105% literal, MIXED)
    rev = [820.0, 905.0, 960.0, 1010.0, 1105.0, 1190.0]
    put(8, "Revenue", rev + [f"={col(5+k)}8*105%" for k in range(1, 5)])   # 2024-27 formula
    put(9, "   % growth", [None] + [f"={col(i)}8/{col(i-1)}8-1" for i in range(1, len(YEARS))])
    # COGS: hardcoded in history (plug: Gross profit derived here); forecast = GrossProfit-Revenue
    cogs = [-500.0, -545.0, -575.0, -600.0, -650.0, -700.0]
    put(11, "- COGS", cogs + [f"={col(6+k)}13-{col(6+k)}8" for k in range(4)])   # 2024-27 = GP-Rev
    # Gross profit: history = Rev+COGS; forecast hardcoded (the plug flips)
    put(13, "= Gross profit",
        [f"={col(i)}8+{col(i)}11" for i in range(6)] + [520.0, 560.0, 600.0, 645.0])
    put(14, "   % margin", [f"={col(i)}13/{col(i)}8" for i in range(len(YEARS))])
    # Opex, EBITDA, D&A, EBIT
    put(16, "- Opex", [-230, -250, -262, -275, -300, -322, -338, -352, -366, -380])
    put(18, "EBITDA", [f"={col(i)}13+{col(i)}16" for i in range(len(YEARS))])
    put(19, "   % margin", [f"={col(i)}18/{col(i)}8" for i in range(len(YEARS))])
    put(20, "- D&A", [-45, -48, -50, -52, -55, -58, -60, -62, -64, -66])
    put(22, "= EBIT", [f"={col(i)}18+{col(i)}20" for i in range(len(YEARS))])
    put(23, "   % margin", [f"={col(i)}22/{col(i)}8" for i in range(len(YEARS))])
    # Net financial (self prior-year carry — Lag test), PBT, Tax (rate driver), Net income
    put(25, "- Net financial", [-18] + [f"={col(i-1)}25" for i in range(1, len(YEARS))])
    put(27, "= PBT", [f"={col(i)}22+{col(i)}25" for i in range(len(YEARS))])
    put(28, "   Tax rate", [0.26] * 6 + [0.25, 0.25, 0.24, 0.24])
    put(29, "- Tax", [f"=-{col(i)}27*{col(i)}28" for i in range(len(YEARS))])
    put(31, "= Net income", [f"={col(i)}27+{col(i)}29" for i in range(len(YEARS))])
    put(32, "   % growth", [None] + [f"={col(i)}31/{col(i-1)}31-1" for i in range(1, len(YEARS))])

    # ---- PER SHARE DATA (the "done right" section: one formula per line) ----
    ws["A35"] = "PER SHARE DATA"
    put(36, "Number of shares (m)", [100.0] * len(YEARS))
    put(37, "EPS", [f"={col(i)}31/{col(i)}36" for i in range(len(YEARS))])
    put(38, "DPS", [0.30, 0.33, 0.35, 0.38, 0.42, 0.46, 0.50, 0.54, 0.58, 0.62])
    put(39, "Payout (%)", [f"={col(i)}38/{col(i)}37" for i in range(len(YEARS))])

    # ---- BALANCE SHEET (a couple lines + a deliberate #REF! forecast cell) ----
    ws["A42"] = "BALANCE SHEET"
    put(43, "Net debt", [310, 295, 280, 260, 240, 215, 190, 160, 130, 95])
    put(44, "Equity", [640, 690, 745, 810, 890, 980, 1080, 1190, 1310, 1440])
    # ND/EBITDA ratio, with a broken 2019 cell (#REF!) to exercise the freeze
    put(45, "Net debt / EBITDA",
        [f"={col(i)}43/{col(i)}18" for i in range(len(YEARS))])
    ws[f"{col(1)}45"] = "=#REF!/#REF!"                      # deliberate broken cell

    out = os.path.join(os.path.dirname(__file__), "Sample Research Model.xlsx")
    wb.save(out)
    inject_cached(out, compute_values())     # give formula cells cached <v> (like a real FactSet file)
    print("wrote", out, "— 10 years, 3 sections, plug-cycle + %-literal + Lag + #REF! + MIXED")
    return out


def compute_values():
    """Evaluate the model in Python so the fixture carries cached values (openpyxl can't
    write formula+cache; real FactSet files always have them). Returns {cellref: number}."""
    n = len(YEARS)
    rev = [820., 905., 960., 1010., 1105., 1190.]
    for _ in range(4):
        rev.append(round(rev[-1] * 1.05, 6))
    opex = [-230, -250, -262, -275, -300, -322, -338, -352, -366, -380]
    da = [-45, -48, -50, -52, -55, -58, -60, -62, -64, -66]
    gp_fc = [520., 560., 600., 645.]
    gp = [rev[i] + (-500, -545, -575, -600, -650, -700)[i] for i in range(6)] + gp_fc
    cogs = [-500., -545., -575., -600., -650., -700.] + [gp[6 + k] - rev[6 + k] for k in range(4)]
    ebitda = [gp[i] + opex[i] for i in range(n)]
    ebit = [ebitda[i] + da[i] for i in range(n)]
    netfin = [-18.] * n
    pbt = [ebit[i] + netfin[i] for i in range(n)]
    taxrate = [0.26] * 6 + [0.25, 0.25, 0.24, 0.24]
    tax = [-pbt[i] * taxrate[i] for i in range(n)]
    ni = [pbt[i] + tax[i] for i in range(n)]
    shares = [100.] * n
    eps = [ni[i] / shares[i] for i in range(n)]
    dps = [0.30, 0.33, 0.35, 0.38, 0.42, 0.46, 0.50, 0.54, 0.58, 0.62]
    payout = [dps[i] / eps[i] for i in range(n)]
    netdebt = [310, 295, 280, 260, 240, 215, 190, 160, 130, 95]
    equity = [640, 690, 745, 810, 890, 980, 1080, 1190, 1310, 1440]
    grow = lambda s: [None] + [s[i] / s[i - 1] - 1 for i in range(1, n)]
    rows = {8: rev, 9: grow(rev), 11: cogs, 13: gp,
            14: [gp[i] / rev[i] for i in range(n)], 16: opex, 18: ebitda,
            19: [ebitda[i] / rev[i] for i in range(n)], 20: da, 22: ebit,
            23: [ebit[i] / rev[i] for i in range(n)], 25: netfin, 27: pbt,
            28: taxrate, 29: tax, 31: ni, 32: grow(ni), 36: shares, 37: eps,
            38: dps, 39: payout, 43: netdebt, 44: equity,
            45: [netdebt[i] / ebitda[i] for i in range(n)]}
    rows[45][1] = None                              # the deliberate #REF! cell has no value
    vals = {5 + 0: None}
    out = {}
    for r, series in rows.items():
        for i, v in enumerate(series):
            if v is not None:
                out[f"{col(i)}{r}"] = round(float(v), 6)
    # year header cells 2020..2027 are =prev+1
    for i in range(2, len(YEARS)):
        out[f"{col(i)}5"] = YEARS[i]
    return out


def inject_cached(path, values):
    """Add a cached <v> to each formula cell so data_only reads return numbers."""
    tmp = path + ".tmp"
    with zipfile.ZipFile(path) as zin:
        names = zin.namelist()
        sheet = next(n for n in names if re.match(r"xl/worksheets/sheet1\.xml", n))
        data = {n: zin.read(n) for n in names}
    xml = data[sheet].decode("utf-8")
    for ref, val in values.items():
        # openpyxl writes formula cells as <c r=".."><f>..</f><v /></c> -> fill the empty <v>
        xml = re.sub(rf'(<c r="{ref}"[^>]*><f[^>]*>.*?</f>)<v ?/>',
                     rf'\1<v>{val}</v>', xml, count=1)
    data[sheet] = xml.encode("utf-8")
    with zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED) as zout:
        for n in names:
            zout.writestr(n, data[n])
    shutil.move(tmp, path)


if __name__ == "__main__":
    build()
