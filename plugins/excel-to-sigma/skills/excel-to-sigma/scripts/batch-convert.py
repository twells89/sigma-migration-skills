#!/usr/bin/env python3
"""
batch-convert.py — fleet converter for a family of broker equity-research models
(the ~850-file case: same house template, "similar but not identical").

Per file: strip FactSet cruft (logged) -> fingerprint the template -> run canonical
inference (infer-canonical-formulas.py) -> map line items to the canonical
chart-of-accounts (coa.json) -> parity gate -> bucket into AUTO_PARITY / NEEDS_REVIEW
/ FAILED. Emits a triage manifest (JSON + CSV) with a rollup and the top unmapped
labels across the batch (which feed COA-alias growth). NO silent truncation.

Default = report-only (recommended calibration mode for the first tranche): inference
+ triage, nothing built. `--post --conn <id> --folder <id>` also builds AUTO_PARITY (and,
with `--post-review`, NEEDS_REVIEW) files via build-research-model.py — READ-ONLY DMs
(inline VALUES), fully API, no UI step (see refs/research-recipes.md).

Usage:
  batch-convert.py <file-or-dir> [<more> ...] [--sheet "FY results"]
                   [--out manifest.json] [--post --conn <id> --folder <id> [--post-review]]
"""
import sys, os, re, json, glob, hashlib, importlib.util
from collections import Counter

HERE = os.path.dirname(os.path.abspath(__file__))


def load_mod(fn, name):
    spec = importlib.util.spec_from_file_location(name, os.path.join(HERE, fn))
    m = importlib.util.module_from_spec(spec); spec.loader.exec_module(m)
    return m


INF = load_mod("infer-canonical-formulas.py", "inf")
COA = json.load(open(os.path.join(HERE, "coa.json")))


def norm(s):
    return re.sub(r"[^a-z0-9 ]", " ", (s or "").lower()).strip()


def alias_index(coa):
    idx = {}
    for acc in coa["accounts"]:
        for a in acc["aliases"]:
            idx[norm(a)] = acc["id"]
    return idx


ALIAS = alias_index(COA)
STABLE_SECTIONS = set(COA["template"]["sections_expected"])


def strip_report(path):
    """Count (never silently drop) the FactSet cruft we ignore."""
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=False, read_only=True)
    named = len(wb.defined_names)
    fdscache = any(s.startswith("__") for s in wb.sheetnames)
    return {"named_ranges": named, "fdscache_sheet": fdscache, "sheets": list(wb.sheetnames)}


def is_detail(l):
    """Ratio / sub-item lines aren't canonical COA accounts (they're carried detail)."""
    n = norm(l["label"])
    return l["kind"] == "ratio" or n.startswith(("ow ", "% ")) or n in ("", ".") or "%" in l["label"][:3]


def fingerprint(plan, sheets, n_mapped):
    """Similarity (not equality) to the house template -> SAME / VARIANT / UNKNOWN."""
    T = COA["template"]; w = T["fingerprint_weights"]
    file_secs = {norm(l["section"]) for l in plan["lines"]}
    exp_secs = {norm(s) for s in T["sections_expected"]}
    sec_score = len(file_secs & exp_secs) / max(len(exp_secs), 1)
    sheet_score = len({norm(s) for s in sheets} & {norm(s) for s in T["sheets_expected"]}) / \
                  max(len(T["sheets_expected"]), 1)
    accountish = [l for l in plan["lines"] if not is_detail(l)]      # headline lines only
    lab_score = n_mapped / max(len(accountish), 1)
    # formula-shape signal: how many derived canonicals look like a COA shape (rough: has refs)
    shp = sum(1 for l in plan["lines"] if l["kind"] in ("derived", "ratio") and l.get("canonical"))
    shp_score = min(shp / 20.0, 1.0)
    score = (w["sections"] * sec_score + w["sheets"] * sheet_score +
             w["label_jaccard"] * lab_score + w["formula_shapes"] * shp_score)
    cls = "SAME" if score >= T["thresholds"]["same"] else \
          ("VARIANT" if score >= T["thresholds"]["variant"] else "UNKNOWN")
    return round(score, 3), cls, {"sections": round(sec_score, 2), "sheets": round(sheet_score, 2),
                                  "labels": round(lab_score, 2)}


def coa_map(plan):
    """Map each line to a canonical account (exact alias, then fuzzy token overlap).
    Unmapped lines are CARRIED (company-specific), never dropped."""
    mapped = {}; unmapped = []
    for l in plan["lines"]:
        n = norm(l["label"])
        acc = ALIAS.get(n)
        if not acc:                                   # fuzzy: alias fully contained / token overlap
            toks = set(n.split())
            best, bestj = None, 0.0
            for a, aid in ALIAS.items():
                at = set(a.split())
                if not at:
                    continue
                j = len(toks & at) / len(toks | at)
                if j > bestj:
                    bestj, best = j, aid
            acc = best if bestj >= 0.6 else None
        if acc:
            mapped[l["row"]] = acc
        else:
            unmapped.append(l["label"])
    return mapped, unmapped


def triage(path, sheet):
    entry = {"file": os.path.basename(path),
             "hash": hashlib.sha1(open(path, "rb").read()).hexdigest()[:12]}
    try:
        strip = strip_report(path)
        entry["stripped"] = strip
        plan = INF.infer(path, sheet=sheet)
    except BaseException as ex:                          # incl. ValueError from axis detection
        entry.update(bucket="FAILED", fingerprint_class="UNKNOWN",
                     reasons=[f"inference error: {ex}"])
        return entry, None
    mapped, unmapped = coa_map(plan)
    sc, cls, det = fingerprint(plan, entry["stripped"]["sheets"], len(mapped))
    s = plan["summary"]; par = plan["parity"]
    n_periods = len(plan["periods"])
    anchors = list(plan["anchors"])
    # a real gap = a HEADLINE line (not a ratio/sub-item) in a stable section that didn't map
    unmapped_stable = [l["label"] for l in plan["lines"]
                       if l["row"] not in mapped and l["section"] in STABLE_SECTIONS
                       and not is_detail(l)]
    reasons = []
    # no-silent-truncation invariants
    if n_periods == 0:
        reasons.append("INVARIANT: no year axis detected")
    if not anchors:
        reasons.append("INVARIANT: no anchor line (revenue/EBIT/net income/EPS) resolved")
    sections_found = {l["section"] for l in plan["lines"]} & STABLE_SECTIONS
    if len(sections_found) < 2:
        reasons.append(f"INVARIANT: only {len(sections_found)} known section(s) found")
    entry.update(sheet=plan["sheet"], fingerprint=sc, fingerprint_class=cls, fingerprint_detail=det,
                 n_periods=n_periods, periods=[p["year"] for p in plan["periods"]][:2] +
                 [plan["periods"][-1]["year"]] if plan["periods"] else [],
                 n_lines=len(plan["lines"]), input=s.get("input", 0), derived=s.get("derived", 0),
                 ratio=s.get("ratio", 0), carried=s.get("carried_plug", 0),
                 needs_review=s.get("needs_review", 0), live=par.get("live", 0),
                 frozen=par.get("frozen", 0), anchors=anchors,
                 n_mapped=len(mapped), n_unmapped=len(unmapped),
                 unmapped_stable=unmapped_stable[:8], reasons=reasons)
    # bucket
    if cls == "UNKNOWN" or any(r.startswith("INVARIANT") for r in reasons):
        entry["bucket"] = "FAILED"
    elif (cls == "SAME" and s.get("needs_review", 0) == 0 and not unmapped_stable):
        entry["bucket"] = "AUTO_PARITY"
    else:
        entry["bucket"] = "NEEDS_REVIEW"
        if cls == "VARIANT":
            reasons.append("VARIANT fingerprint")
        if s.get("needs_review", 0):
            reasons.append(f"{s['needs_review']} line(s) flagged in inference")
        if unmapped_stable:
            reasons.append(f"{len(unmapped_stable)} unmapped line(s) in stable sections (carried)")
    return entry, plan


def expand(paths):
    out = []
    for p in paths:
        if os.path.isdir(p):
            out += sorted(glob.glob(os.path.join(p, "*.xlsx")))
        else:
            out += sorted(glob.glob(p)) if any(c in p for c in "*?[") else [p]
    return [p for p in out if p.lower().endswith(".xlsx") and "~$" not in p]


def main():
    args = [a for a in sys.argv[1:] if not a.startswith("--")]
    opt = {a.split("=")[0]: (a.split("=", 1)[1] if "=" in a else True)
           for a in sys.argv[1:] if a.startswith("--")}
    files = expand(args)
    if not files:
        sys.exit(__doc__)
    sheet = opt.get("--sheet")
    manifest = []; plans = {}
    for f in files:
        entry, plan = triage(f, sheet)
        manifest.append(entry); plans[entry["file"]] = plan
        print(f"  [{entry['bucket']:12s}] {entry['file'][:48]:48s} "
              f"fp={entry.get('fingerprint','-')}/{entry.get('fingerprint_class','-')} "
              f"periods={entry.get('n_periods','-')} live={entry.get('live','-')} "
              f"frozen={entry.get('frozen','-')} unmapped={entry.get('n_unmapped','-')}")

    buckets = Counter(e["bucket"] for e in manifest)
    top_unmapped = Counter()
    for e in manifest:
        for lab in (e.get("unmapped_stable") or []):
            top_unmapped[lab] += 1
    print(f"\nROLLUP over {len(manifest)} file(s): " +
          "  ".join(f"{k}={v}" for k, v in buckets.most_common()))
    if top_unmapped:
        print("Top unmapped stable-section labels (grow coa.json aliases):")
        for lab, n in top_unmapped.most_common(10):
            print(f"   {n:3d}  {lab}")

    out = opt.get("--out")
    if out and out is not True:
        json.dump({"buckets": dict(buckets), "top_unmapped": top_unmapped.most_common(20),
                   "files": manifest}, open(out, "w"), indent=1)
        csvp = os.path.splitext(out)[0] + ".csv"
        import csv
        keys = ["file", "bucket", "fingerprint", "fingerprint_class", "n_periods", "n_lines",
                "input", "derived", "ratio", "carried", "needs_review", "live", "frozen",
                "n_mapped", "n_unmapped"]
        with open(csvp, "w", newline="") as fh:
            w = csv.DictWriter(fh, fieldnames=keys, extrasaction="ignore"); w.writeheader()
            for e in manifest:
                w.writerow(e)
        print(f"\nmanifest: {out}  +  {csvp}")

    # optional live build for AUTO_PARITY (+ NEEDS_REVIEW with --post-review)
    if "--post" in opt:
        BUILD = load_mod("build-research-model.py", "brm")
        conn, folder = opt.get("--conn"), opt.get("--folder")
        want = {"AUTO_PARITY"} | ({"NEEDS_REVIEW"} if "--post-review" in opt else set())
        e = BUILD.env(); base = e["SIGMA_BASE_URL"]; tok = BUILD.token(e)
        for entry in manifest:
            if entry["bucket"] not in want:
                continue
            plan = plans[entry["file"]]
            name = os.path.splitext(entry["file"])[0][:56] + " (auto)"
            dl = [ln for ln in plan["lines"] if ln["typed"]]
            dm = BUILD.api(base, tok, "POST", "/v2/dataModels/spec",
                           BUILD.build_dm(plan, conn, folder, name)[0])
            full = BUILD.api(base, tok, "GET", f"/v2/dataModels/{dm['dataModelId']}/spec")
            # Released workbook code_rep: wrap nested draft + layout last before POST.
            wb_body = BUILD.workbook_wire.wire_workbook(
                BUILD.build_wb(plan, dm["dataModelId"], BUILD.map_dm(full), dl, folder, name))
            wb = BUILD.api(base, tok, "POST", "/v2/workbooks/spec", wb_body)
            print(f"   built {entry['file']}: wb {wb.get('workbookId')}")


if __name__ == "__main__":
    main()
