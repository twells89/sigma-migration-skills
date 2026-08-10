#!/usr/bin/env python3
"""Phase 1 — extract a formal Table to a paste-ready input-table CSV.

Emits ONLY the data-entry columns with UPPER_SNAKE_CASE headers (so a clipboard
paste into the Sigma input table aligns 1:1), excludes any system-style columns,
and writes Excel dates as ISO yyyy-mm-dd. Also prints suggested input-table column
specs (name:type) for build-input-table-wb.py.

Usage:
    python xlsx-to-input-csv.py "Sample Forecast.xlsx" --table tblForecast --out forecast_input_paste.csv
"""
import argparse
import csv
import datetime as dt
import re
from openpyxl import load_workbook

SYSTEM_COLS = {"ID", "ROW_ID", "CREATED_AT", "CREATED_BY",
               "UPDATED_AT", "UPDATED_BY", "LAST_UPDATED_AT", "LAST_UPDATED_BY"}


def snake(name):
    s = str(name)
    # split camelCase / PascalCase boundaries: SubBranch -> Sub_Branch
    s = re.sub(r"(?<=[a-z0-9])(?=[A-Z])", "_", s)
    s = re.sub(r"[^0-9A-Za-z]+", "_", s).strip("_").upper()
    return s


def infer_type(values):
    seen = set()
    for v in values:
        if v is None or v == "":
            continue
        if isinstance(v, bool):
            seen.add("checkbox")
        elif isinstance(v, (dt.datetime, dt.date)):
            seen.add("datetime")
        elif isinstance(v, (int, float)):
            seen.add("number")
        else:
            seen.add("text")
    if not seen:
        return "text"
    if seen == {"number"}:
        return "number"
    if seen <= {"datetime"}:
        return "datetime"
    if seen == {"checkbox"}:
        return "checkbox"
    return "text"


def find_table(wb, table_name):
    for ws in wb.worksheets:
        if table_name in ws.tables:
            return ws, ws.tables[table_name]
    raise SystemExit(f"table '{table_name}' not found. "
                     f"available: {[n for ws in wb.worksheets for n in ws.tables]}")


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("xlsx")
    ap.add_argument("--table", required=True)
    ap.add_argument("--out", required=True)
    args = ap.parse_args()

    wb = load_workbook(args.xlsx, data_only=True)
    ws, tbl = find_table(wb, args.table)
    cells = ws[tbl.ref]
    raw_headers = [c.value for c in cells[0]]
    body = cells[1:]

    keep_idx, out_headers, col_values = [], [], []
    for i, h in enumerate(raw_headers):
        hsnake = snake(h)
        if hsnake in SYSTEM_COLS:
            continue
        keep_idx.append(i)
        out_headers.append(hsnake)
        col_values.append([row[i].value for row in body])

    def fmt(v):
        if isinstance(v, (dt.datetime, dt.date)):
            return v.strftime("%Y-%m-%d")
        return v

    with open(args.out, "w", newline="") as f:
        w = csv.writer(f)
        w.writerow(out_headers)
        for row in body:
            w.writerow([fmt(row[i].value) for i in keep_idx])

    print(f"wrote {args.out}: {len(body)} rows, {len(out_headers)} entry columns")
    print(f"  headers: {','.join(out_headers)}")
    specs = [f"{h}:{infer_type(col_values[j])}" for j, h in enumerate(out_headers)]
    print("\nsuggested --columns for build-input-table-wb.py:")
    print("  " + ",".join(specs))


if __name__ == "__main__":
    main()
