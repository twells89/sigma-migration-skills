#!/usr/bin/env python3
"""Phase 4 — POST a data model that reads FROM the input-table warehouse view.

This is the source-swap target: a fact element sourced from
`SELECT … FROM <view>`, plus (optionally) a category dimension lifted from a
formal Table in the Excel file and a flattened *_REPORT element (in-SQL join) for
section rollups.

Auth: reads ~/.sigma-migration/env. Verify parity afterward by querying the
report element (Sigma MCP) grouped by the rollup dimension.

Usage:
    python build-dm-on-view.py --view SIGMA_WRITE_DB.SIGMA_WRITE.FORECAST_ENTRY \
        --connection cb2f5180-… \
        --fact-cols REGION,BRANCH,SUB_BRANCH,MONTH_DATE,CATEGORY_CODE,FORECAST_AMOUNT \
        [--categories-from "Sample Forecast.xlsx" --categories-table tblCategory \
         --join-key CATEGORY_CODE]
"""
import argparse
import json
import os
import urllib.request
import urllib.parse


def token():
    env = {}
    with open(os.path.expanduser("~/.sigma-migration/env")) as f:
        for line in f:
            line = line.strip().replace("export ", "")
            if "=" in line:
                k, v = line.split("=", 1)
                env[k] = v.strip().strip('"').strip("'")
    data = urllib.parse.urlencode({
        "grant_type": "client_credentials",
        "client_id": env["SIGMA_CLIENT_ID"],
        "client_secret": env["SIGMA_CLIENT_SECRET"],
    }).encode()
    req = urllib.request.Request(env["SIGMA_BASE_URL"] + "/v2/auth/token", data=data,
                                 headers={"Content-Type": "application/x-www-form-urlencoded"})
    return env["SIGMA_BASE_URL"], json.load(urllib.request.urlopen(req))["access_token"]


def home_folder(base, tok):
    req = urllib.request.Request(base + "/v2/whoami",
                                 headers={"Authorization": f"Bearer {tok}"})
    uid = json.load(urllib.request.urlopen(req))["userId"]
    req = urllib.request.Request(base + f"/v2/members/{uid}",
                                 headers={"Authorization": f"Bearer {tok}"})
    return json.load(urllib.request.urlopen(req)).get("homeFolderId")


def title(alias):
    return alias.replace("_", " ").title()


def cols(aliases, prefix):
    # SQL-element columns MUST use the literal [Custom SQL/ALIAS] prefix
    return [{"id": f"{prefix}{i}", "formula": f"[Custom SQL/{a}]", "name": title(a)}
            for i, a in enumerate(aliases, 1)]


def category_values(xlsx, table):
    """Lift a formal Table into a VALUES list (codes assumed numeric-or-text)."""
    from openpyxl import load_workbook
    wb = load_workbook(xlsx, data_only=True)
    for ws in wb.worksheets:
        if table in ws.tables:
            cells = ws[ws.tables[table].ref]
            headers = [str(c.value) for c in cells[0]]
            rows = []
            for r in cells[1:]:
                vals = []
                for c in r:
                    v = c.value
                    if isinstance(v, (int, float)):
                        vals.append(str(v))
                    else:
                        vals.append("'" + str(v).replace("'", "''") + "'")
                rows.append("(" + ",".join(vals) + ")")
            return headers, ",\n".join(rows)
    raise SystemExit(f"category table '{table}' not found in {xlsx}")


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--view", required=True, help="fully qualified db.schema.view")
    ap.add_argument("--connection", required=True)
    ap.add_argument("--fact-cols", required=True, help="comma list of view column names")
    ap.add_argument("--folder", default=None)
    ap.add_argument("--name", default="Forecast (Input Table)")
    ap.add_argument("--categories-from", default=None)
    ap.add_argument("--categories-table", default="tblCategory")
    ap.add_argument("--join-key", default="CATEGORY_CODE")
    args = ap.parse_args()

    base, tok = token()
    folder = args.folder or home_folder(base, tok)
    fact_cols = [c.strip() for c in args.fact_cols.split(",")]

    fact_sql = f"select {', '.join(fact_cols)} from {args.view}"
    elements = [{
        "id": "fact", "kind": "table",
        "source": {"kind": "sql", "connectionId": args.connection, "statement": fact_sql},
        "columns": cols(fact_cols, "fc"),
    }]

    if args.categories_from:
        chdr, cvals = category_values(args.categories_from, args.categories_table)
        cat_aliases = [h.replace(" ", "_").upper() for h in chdr]
        cat_sql = ("select " +
                   ", ".join(f"column{i+1} as {a}" for i, a in enumerate(cat_aliases)) +
                   f" from values\n{cvals}")
        elements.append({
            "id": "category", "kind": "table",
            "source": {"kind": "sql", "connectionId": args.connection, "statement": cat_sql},
            "columns": cols(cat_aliases, "ca"),
        })
        # flattened report (in-SQL join — avoids the cross-element rollup error)
        join_cat_cols = [a for a in cat_aliases if a != args.join_key]
        report_sql = (
            f"with f as ({fact_sql}),\n"
            f"c as ({cat_sql})\n"
            f"select " + ", ".join(f"f.{a}" for a in fact_cols) + ", " +
            ", ".join(f"c.{a}" for a in join_cat_cols) +
            f" from f join c on f.{args.join_key} = c.{args.join_key}")
        elements.append({
            "id": "report", "kind": "table",
            "source": {"kind": "sql", "connectionId": args.connection, "statement": report_sql},
            "columns": cols(fact_cols + join_cat_cols, "fr"),
        })

    spec = {"name": args.name, "schemaVersion": 1, "folderId": folder,
            "pages": [{"id": "model", "name": "Model", "elements": elements}]}

    body = json.dumps(spec).encode()
    req = urllib.request.Request(base + "/v2/dataModels/spec", data=body, method="POST",
                                 headers={"Authorization": f"Bearer {tok}",
                                          "Content-Type": "application/json"})
    resp = json.load(urllib.request.urlopen(req))
    print("dataModelId:", resp.get("dataModelId"))
    print("url:", resp.get("url"))
    print("\nNEXT: `describe` the model via the Sigma MCP (element + column IDs are "
          "reassigned), then query the report element grouped by your rollup "
          "dimension and compare to parity targets.")


if __name__ == "__main__":
    main()
